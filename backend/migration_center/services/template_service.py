"""Downloadable CSV/XLSX templates and error/reconciliation reports."""

from __future__ import annotations

import csv
import io

from migration_center.datasets import get_dataset
from migration_center.models import MigrationBatch, StagingRowStatus


def template_csv(dataset_key: str) -> tuple[str, bytes]:
    dataset = get_dataset(dataset_key)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([spec.label for spec in dataset.fields])
    return f"{dataset.key}_template.csv", buffer.getvalue().encode("utf-8-sig")


def template_xlsx(dataset_key: str) -> tuple[str, bytes]:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    dataset = get_dataset(dataset_key)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = dataset.label[:31]
    bold = Font(bold=True)
    for column, spec in enumerate(dataset.fields, start=1):
        cell = sheet.cell(row=1, column=column, value=spec.label)
        cell.font = bold
        hints = []
        if spec.required:
            hints.append("required")
        if spec.kind == "date":
            hints.append("YYYY-MM-DD")
        if spec.kind == "decimal":
            hints.append("number")
        if spec.choices:
            hints.append("/".join(spec.choices))
        sheet.cell(row=2, column=column, value=f"({', '.join(hints)})" if hints else "")
    stream = io.BytesIO()
    workbook.save(stream)
    return f"{dataset.key}_template.xlsx", stream.getvalue()


def error_report_csv(batch: MigrationBatch) -> tuple[str, bytes]:
    dataset = get_dataset(batch.dataset_key)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Row", "Status", "Errors", "Warnings", "Import Error", *[s.label for s in dataset.fields]])
    problem_rows = batch.rows.filter(
        status__in=[StagingRowStatus.ERROR, StagingRowStatus.FAILED, StagingRowStatus.WARNING, StagingRowStatus.DUPLICATE]
    ).order_by("row_number")
    for row in problem_rows.iterator(chunk_size=500):
        writer.writerow([
            row.row_number, row.status,
            " | ".join(row.errors or []), " | ".join(row.warnings or []), row.import_error,
            *[str(row.mapped_data.get(s.key, "") or row.raw_data.get((batch.mapping or {}).get(s.key, ""), "")) for s in dataset.fields],
        ])
    return f"{batch.batch_number}_errors.csv", buffer.getvalue().encode("utf-8-sig")


def reconciliation_report_csv(batch: MigrationBatch) -> tuple[str, bytes]:
    snapshot = batch.reconciliation_snapshot or {}
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Batch", "Dataset", "Expected Total", "Imported Total", "Difference", "Matched", "Imported Rows", "Failed Rows", "Skipped Rows"])
    writer.writerow([
        batch.batch_number, batch.dataset_key,
        snapshot.get("expected_total"), snapshot.get("imported_total"), snapshot.get("difference"),
        snapshot.get("matched"), batch.imported_rows, batch.failed_rows, batch.skipped_rows,
    ])
    return f"{batch.batch_number}_reconciliation.csv", buffer.getvalue().encode("utf-8-sig")
