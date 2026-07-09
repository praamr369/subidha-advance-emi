"""File parsing for the Migration Center — CSV and XLSX, streamed in chunks."""

from __future__ import annotations

import csv
import io
from typing import Any, Iterator


class FileParseError(ValueError):
    pass


def _read_bytes(uploaded_file: Any) -> bytes:
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    raw = uploaded_file.read()
    return raw if isinstance(raw, bytes) else str(raw).encode("utf-8")


def parse_csv(raw: bytes) -> tuple[list[str], Iterator[dict[str, str]]]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    headers = [h for h in (reader.fieldnames or []) if h is not None]
    if not headers:
        raise FileParseError("CSV file has no header row.")

    def rows() -> Iterator[dict[str, str]]:
        for row in reader:
            yield {k: (v or "").strip() for k, v in row.items() if k is not None}

    return headers, rows()


def parse_xlsx(raw: bytes) -> tuple[list[str], Iterator[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise FileParseError("XLSX support requires the openpyxl package.") from exc
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise FileParseError(f"Could not open XLSX file: {exc}") from exc
    sheet = workbook.active
    if sheet is None:
        raise FileParseError("XLSX file has no active worksheet.")
    row_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        raise FileParseError("XLSX file is empty.")
    headers = [str(cell).strip() for cell in header_row if cell is not None and str(cell).strip()]
    if not headers:
        raise FileParseError("XLSX file has no header row.")

    def rows() -> Iterator[dict[str, str]]:
        try:
            for values in row_iter:
                row: dict[str, str] = {}
                for idx, header in enumerate(headers):
                    value = values[idx] if idx < len(values) else None
                    row[header] = "" if value is None else str(value).strip()
                if any(v for v in row.values()):
                    yield row
        finally:
            workbook.close()

    return headers, rows()


def parse_upload(uploaded_file: Any, filename: str) -> tuple[list[str], Iterator[dict[str, str]], bytes]:
    """Return (headers, row iterator, raw bytes) for a CSV/XLSX upload."""
    raw = _read_bytes(uploaded_file)
    lowered = (filename or "").lower()
    if lowered.endswith(".xlsx") or lowered.endswith(".xlsm"):
        headers, rows = parse_xlsx(raw)
    elif lowered.endswith(".csv") or lowered.endswith(".txt"):
        headers, rows = parse_csv(raw)
    else:
        # Sniff: XLSX files are ZIP archives (PK header).
        if raw[:2] == b"PK":
            headers, rows = parse_xlsx(raw)
        else:
            headers, rows = parse_csv(raw)
    return headers, rows, raw
